from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class LoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

    def login(self, username, password):
        username_input = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "txtUsername")))
        username_input.clear()
        username_input.send_keys(username)

        password_input = self.driver.find_element(By.ID, "txtPassword")
        password_input.clear()
        password_input.send_keys(password)

        login_button = self.driver.find_element(By.ID, "btnLogin")
        login_button.click()

    def is_login_successful(self):
        return EC.presence_of_element_located((By.ID, "welcome"))(self.driver)


import datetime
import openpyxl
from selenium import webdriver
from page_objects import LoginPage
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class TestLogin:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.login_page = LoginPage(self.driver)

    def teardown_method(self):
        self.driver.quit()

    def test_login_with_excel_data(self):
        workbook = openpyxl.load_workbook("login_data.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            test_id, username, password, _, _, _, _ = row

            self.login_page.login(username, password)

            if self.login_page.is_login_successful():
                sheet.cell(row=test_id, column=7, value="Pass")
            else:
                sheet.cell(row=test_id, column=7, value="Fail")

        workbook.save("login_data.xlsx")

if __name__ == "__main__":
    pytest.main()